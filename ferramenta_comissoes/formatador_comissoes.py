# -*- coding: utf-8 -*-
"""
Formatador de Comissões (protótipo).

Fluxo:
  1. No topo da janela: Ano e Mês, compartilhados por todas as abas. Trocar
     o período aqui atualiza o que aparece nas abas "Configurar E-mails" e
     "Enviar".
  2. Aba "Processar": escolher as pastas (Downloads e saída), com padrão
     salvo. Ao clicar em "Processar", varre a pasta de Downloads por
     arquivos no padrão "comissoes_AAAA_MM_<Equipe>.xlsx", formata TODAS as
     abas de cada um e salva como "Comissões OTE AAAAMM - <Nome> vN.xlsx" na
     pasta de saída (versão anterior vai para "V Olds").
  3. Aba "Configurar E-mails": destinatário (Para) e Cc por equipe, salvos
     entre execuções. Só lista equipes com arquivo baixado ou já processado
     para o Ano/Mês selecionado no topo.
  4. Aba "Enviar": lista os arquivos já gerados para o Ano/Mês selecionado,
     permite escolher quais enviar, o assunto e o corpo (com placeholders
     {equipe}/{mes}/{ano}), e dispara um e-mail por arquivo via Outlook,
     como rascunho para revisão ou envio direto, conforme escolhido.

Requer: Python 3.9+, openpyxl (pip install openpyxl). Envio de e-mail usa smtplib da biblioteca padrão.
Gerar .exe: ver README.md
"""

import os
import sys
import re
import json
import shutil

# No .exe (PyInstaller onefile) aponta o Tcl/Tk para os dados empacotados;
# sem isso o _tkinter/Tk não encontra as libs e a janela não abre.
if hasattr(sys, "_MEIPASS"):
    os.environ.setdefault("TCL_LIBRARY", os.path.join(sys._MEIPASS, "tcl8.6"))
    os.environ.setdefault("TK_LIBRARY", os.path.join(sys._MEIPASS, "tk8.6"))

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────── Configuração ────────────────────────────────────

APP_NAME = "FormatadorComissoes"
CONFIG_PATH = os.path.join(
    os.environ.get("APPDATA", os.path.expanduser("~")), APP_NAME, "config.json"
)

DEFAULT_DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")
DEFAULT_OUTPUT = (
    r"C:\Users\Higor.Nocetti\OneDrive - S3eng - Tecnologia Aplicada à Engenharia LTDA"
    r"\Bonificação e Comissão"
)
OLDS_SUBDIR = "V Olds"

# Mapa explícito: slug no nome do arquivo (após "comissoes_AAAA_MM_") -> nome final.
# Ajuste/complete conforme suas equipes. Slugs não listados caem no fallback
# (troca "_" por espaço).
EQUIPE_NOME = {
    "Ares": "Ares",
    "B2B_Construtora": "B2B Construtoras",
    "B2B_Escritório": "B2B Escritórios",
    "FSB": "FSB",
    "Farmer": "Farmer",
    "GD": "GD",
    "Governo": "B2G",   # no painel é "Governo"; para o DP o nome final é "B2G"
    "Saving": "Saving",
    "Cancelamento": "Cancelamento",
}

# Grupos de envio: equipes com arquivos separados (nomes e versões próprias),
# mas que devem sair num único e-mail, com mais de um anexo, para o mesmo
# destinatário. Equipes fora de qualquer grupo formam um grupo próprio, com
# elas mesmas, usando o slug como chave de configuração (compatível com o
# que já estava salvo antes de existir agrupamento).
GRUPOS_ENVIO = {
    "Saving e Cancelamento": ["Saving", "Cancelamento"],
}

# Categorias (linha 2) que recebem tratamento especial.
CAT_TOTALIZACAO = "Totalizações"
CAT_ACELERADORES = "Aceleradores Forma de Pagamento"

MESES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}
MES_NUM = {v.lower(): k for k, v in MESES.items()}

DEFAULT_ASSUNTO = "Comissões {equipe} - {mes}/{ano} para validação"
DEFAULT_CORPO = (
    "Olá,\n\n"
    "Segue em anexo a planilha de comissões de {equipe} referente a {mes}/{ano} "
    "para validação.\n\n"
    "Qualquer divergência, me avise.\n\n"
    "Obrigado!"
)


# ─────────────────────────── Persistência de config ──────────────────────────

def load_config():
    try:
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(update):
    """Mescla `update` no config já salvo (não apaga chaves de outras abas)."""
    cfg = load_config()
    cfg.update(update)
    try:
        os.makedirs(os.path.dirname(CONFIG_PATH), exist_ok=True)
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ─────────────────────────── Regras de nome ──────────────────────────────────

def nome_equipe(slug):
    return EQUIPE_NOME.get(slug, slug.replace("_", " "))


def slug_da_equipe(nome_arquivo, ano, mes):
    """De 'comissoes_2026_06_B2B_Escritórios (1).xlsx' -> 'B2B_Escritórios'."""
    nome = nome_arquivo
    if nome.lower().endswith(".xlsx"):
        nome = nome[:-5]
    nome = re.sub(r"\s*\(\d+\)$", "", nome)          # remove sufixo " (1)"
    prefixo = f"comissoes_{ano}_{mes:02d}_"
    if nome.startswith(prefixo):
        nome = nome[len(prefixo):]
    return nome


def _base_sem_duplicata(nome_arquivo):
    """'arq (1).xlsx' e 'arq.xlsx' compartilham a mesma base 'arq.xlsx'."""
    return re.sub(r"\s*\(\d+\)(?=\.xlsx$)", "", nome_arquivo, flags=re.IGNORECASE)


# ─────────────────────────── Versão / movimentação ───────────────────────────

def _regex_versao(yyyymm, display):
    return re.compile(
        r"^Comissões OTE " + re.escape(yyyymm) + r" - " + re.escape(display)
        + r" v(\d+)\.xlsx$"
    )


def versoes_existentes(output_dir, yyyymm, display):
    pat = _regex_versao(yyyymm, display)
    achados = {}
    if os.path.isdir(output_dir):
        for f in os.listdir(output_dir):
            m = pat.match(f)
            if m:
                achados[int(m.group(1))] = f
    return achados


def proxima_versao(output_dir, yyyymm, display):
    # Considera a pasta atual E a subpasta 'V Olds' para o número sempre subir,
    # mesmo que a versão anterior já tenha sido movida.
    atual = versoes_existentes(output_dir, yyyymm, display)
    olds = versoes_existentes(os.path.join(output_dir, OLDS_SUBDIR), yyyymm, display)
    todas = set(atual) | set(olds)
    return (max(todas) + 1) if todas else 1


def arquivo_mais_recente(output_dir, yyyymm, display):
    """(caminho, versao) do arquivo final mais recente já gerado, ou None."""
    ex = versoes_existentes(output_dir, yyyymm, display)
    if not ex:
        return None
    versao = max(ex)
    return os.path.join(output_dir, ex[versao]), versao


def mover_anteriores(output_dir, yyyymm, display, log):
    """Move as versões já existentes desta equipe/mês para 'V Olds'."""
    ex = versoes_existentes(output_dir, yyyymm, display)
    if not ex:
        return
    olds = os.path.join(output_dir, OLDS_SUBDIR)
    os.makedirs(olds, exist_ok=True)
    for _v, fname in sorted(ex.items()):
        src = os.path.join(output_dir, fname)
        dst = os.path.join(olds, fname)
        if os.path.exists(dst):
            os.remove(dst)
        shutil.move(src, dst)
        log(f"   → versão anterior movida para '{OLDS_SUBDIR}': {fname}")


# ─────────────────────────── Formatação da planilha ──────────────────────────

def formatar_planilha(ws):
    max_col = ws.max_column
    max_row = ws.max_row
    if max_col < 1 or max_row < 2:
        return  # nada a formatar

    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    cats = [ws.cell(row=2, column=c).value for c in range(1, max_col + 1)]

    # 1) Coluna "Mês" -> número (dados originais estão nas linhas 3..max_row)
    for c in range(1, max_col + 1):
        if str(headers[c - 1]).strip().lower() in ("mês", "mes"):
            for r in range(3, max_row + 1):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                num = MES_NUM.get(str(val).strip().lower())
                if num:
                    ws.cell(row=r, column=c).value = num

    # 2) Remover a linha de categorias (linha 2). Header vira 1, dados 2..
    ws.delete_rows(2, 1)
    new_max_row = ws.max_row

    # 3) Grupos: sequências consecutivas de mesma categoria
    grupos = []
    start = 1
    for c in range(2, max_col + 1):
        if cats[c - 1] != cats[start - 1]:
            grupos.append((cats[start - 1], start, c - 1))
            start = c
    grupos.append((cats[start - 1], start, max_col))

    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="thick", color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # 4) Fonte + alinhamento + ocultar aceleradores
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        cat = cats[c - 1]
        if cat == CAT_ACELERADORES:
            ws.column_dimensions[col_letter].hidden = True
        cor = "FF0000" if cat == CAT_TOTALIZACAO else "000000"
        for r in range(1, new_max_row + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.font = Font(bold=(r == 1), color=cor)

    # 5) Bordas: caixa grossa por grupo, grade fina no interior
    for _cat, sc, ec in grupos:
        for r in range(1, new_max_row + 1):
            for c in range(sc, ec + 1):
                ws.cell(row=r, column=c).border = Border(
                    left=thick if c == sc else thin,
                    right=thick if c == ec else thin,
                    top=thick if r == 1 else thin,
                    bottom=thick if r == new_max_row else thin,
                )

    # 6) Largura aproximada ao conteúdo (ignora colunas ocultas)
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        if ws.column_dimensions[col_letter].hidden:
            continue
        maxlen = 0
        for r in range(1, new_max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[col_letter].width = max(8, min(maxlen + 3, 60))


# ─────────────────────────── Processamento de arquivos ───────────────────────

def processar_arquivo(caminho, ano, mes, output_dir, log):
    slug = slug_da_equipe(os.path.basename(caminho), ano, mes)
    display = nome_equipe(slug)
    yyyymm = f"{ano}{mes:02d}"

    wb = openpyxl.load_workbook(caminho)
    for ws in wb.worksheets:
        formatar_planilha(ws)

    os.makedirs(output_dir, exist_ok=True)
    versao = proxima_versao(output_dir, yyyymm, display)   # calcula ANTES de mover
    mover_anteriores(output_dir, yyyymm, display, log)     # move a anterior p/ V Olds
    nome_final = f"Comissões OTE {yyyymm} - {display} v{versao}.xlsx"
    wb.save(os.path.join(output_dir, nome_final))
    log(f"✓ {os.path.basename(caminho)}  →  {nome_final}")


def processar(ano, mes, downloads_dir, output_dir, log, escolher_dup):
    if not os.path.isdir(downloads_dir):
        log(f"Pasta de Downloads não encontrada: {downloads_dir}")
        return
    prefixo = f"comissoes_{ano}_{mes:02d}_"
    candidatos = [
        f for f in os.listdir(downloads_dir)
        if f.startswith(prefixo) and f.lower().endswith(".xlsx")
    ]
    # avisa se houver CSVs (export sem openpyxl) que serão ignorados
    csvs = [f for f in os.listdir(downloads_dir)
            if f.startswith(prefixo) and f.lower().endswith(".csv")]
    if csvs:
        log(f"Atenção: {len(csvs)} arquivo(s) .csv ignorado(s) (a ferramenta usa .xlsx).")

    if not candidatos:
        log(f"Nenhum .xlsx encontrado com o prefixo '{prefixo}' em {downloads_dir}")
        return

    grupos_dup = {}
    for f in candidatos:
        grupos_dup.setdefault(_base_sem_duplicata(f), []).append(f)

    log(f"{len(grupos_dup)} equipe(s) encontrada(s).")
    for base, arquivos in sorted(grupos_dup.items()):
        if len(arquivos) > 1:
            escolhido = escolher_dup(base, sorted(arquivos))
            if not escolhido:
                log(f"- Pulado (nenhum selecionado): {base}")
                continue
        else:
            escolhido = arquivos[0]
        try:
            processar_arquivo(os.path.join(downloads_dir, escolhido), ano, mes, output_dir, log)
            for f_del in arquivos:
                fp = os.path.join(downloads_dir, f_del)
                try:
                    os.remove(fp)
                    log(f"   🗑 Removido de Downloads: {f_del}")
                except Exception as e_del:
                    log(f"   ⚠ Não foi possível remover de Downloads: {f_del} ({e_del})")
        except Exception as e:
            log(f"✗ ERRO em {escolhido}: {e}")
    log("Concluído.")


# ─────────────────────────── Envio por e-mail (Outlook) ──────────────────────

def _grupos_completos():
    """{chave_config: (nome_exibido, [slugs])}.

    Um grupo de GRUPOS_ENVIO usa o próprio nome do grupo como chave. Uma
    equipe fora de qualquer grupo forma um grupo com ela mesma, usando o seu
    slug como chave, a mesma chave de sempre, para não perder e-mail que já
    tenha sido salvo antes de existir agrupamento."""
    grupos = {nome: (nome, list(slugs)) for nome, slugs in GRUPOS_ENVIO.items()}
    agrupadas = {slug for slugs in GRUPOS_ENVIO.values() for slug in slugs}
    for slug, display in EQUIPE_NOME.items():
        if slug not in agrupadas:
            grupos[slug] = (display, [slug])
    return grupos


def grupos_do_mes(downloads_dir, output_dir, ano, mes):
    """[(chave_config, nome_exibido, [slugs])] dos grupos com pelo menos um
    arquivo (baixado ou já processado) para este Ano/Mês, ordenado por nome.
    Um grupo sem nenhum arquivo de nenhuma das suas equipes não aparece
    (ex.: Ares não aparece em um mês sem exportação de Ares)."""
    prefixo = f"comissoes_{ano}_{mes:02d}_"
    yyyymm = f"{ano}{mes:02d}"

    slugs_baixados = set()
    if os.path.isdir(downloads_dir):
        for f in os.listdir(downloads_dir):
            if f.startswith(prefixo) and f.lower().endswith(".xlsx"):
                slugs_baixados.add(slug_da_equipe(f, ano, mes))

    def _slug_ativo(slug):
        display = nome_equipe(slug)
        return slug in slugs_baixados or bool(arquivo_mais_recente(output_dir, yyyymm, display))

    ativos = [
        (chave, nome, slugs) for chave, (nome, slugs) in _grupos_completos().items()
        if any(_slug_ativo(s) for s in slugs)
    ]
    return sorted(ativos, key=lambda item: item[1])


def montar_grupos_envio(output_dir, ano, mes):
    """[{chave, nome, arquivos: [{slug, display, caminho, versao}]}] só com
    os grupos que tenham pelo menos um arquivo já processado neste Ano/Mês.
    Um grupo com duas equipes (ex.: Saving e Cancelamento) reúne os arquivos
    de cada uma que já exista; a que ainda não foi processada simplesmente
    não entra na lista de anexos."""
    yyyymm = f"{ano}{mes:02d}"
    resultado = []
    for chave, (nome, slugs) in _grupos_completos().items():
        arquivos = []
        for slug in slugs:
            display = nome_equipe(slug)
            achado = arquivo_mais_recente(output_dir, yyyymm, display)
            if achado:
                caminho, versao = achado
                arquivos.append({"slug": slug, "display": display, "caminho": caminho, "versao": versao})
        if arquivos:
            resultado.append({"chave": chave, "nome": nome, "arquivos": arquivos})
    return sorted(resultado, key=lambda g: g["nome"])


def resolver_template(texto, display, mes, ano):
    return (
        texto.replace("{equipe}", display)
             .replace("{mes}", MESES.get(mes, str(mes)))
             .replace("{ano}", str(ano))
    )


def _construir_mensagem(de, para, cc, assunto, corpo, caminhos_anexos):
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    msg = MIMEMultipart()
    msg["From"] = de
    msg["To"] = para
    if cc:
        msg["Cc"] = cc
    msg["Subject"] = assunto
    msg.attach(MIMEText(corpo, "plain", "utf-8"))
    for caminho in caminhos_anexos:
        with open(caminho, "rb") as f:
            parte = MIMEBase("application", "octet-stream")
            parte.set_payload(f.read())
        encoders.encode_base64(parte)
        parte.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(caminho)}"',
        )
        msg.attach(parte)
    return msg


def enviar_email_smtp(conta, senha_app, para, cc, assunto, corpo, caminhos_anexos):
    import smtplib
    import ssl

    msg = _construir_mensagem(conta, para, cc, assunto, corpo, caminhos_anexos)
    destinatarios = [e.strip() for e in para.split(";") if e.strip()]
    if cc:
        destinatarios += [e.strip() for e in cc.split(";") if e.strip()]
    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx) as smtp:
        smtp.login(conta, senha_app)
        smtp.sendmail(conta, destinatarios, msg.as_bytes())


def executar_envio(selecionados, cfg_emails, cfg_smtp, assunto_tmpl, corpo_tmpl,
                   ano, mes, modo_rascunho, log, confirmar_fn=None):
    """Envia um e-mail por grupo selecionado via Gmail SMTP. Retorna (ok, falhas)."""
    conta = cfg_smtp.get("conta", "").strip()
    senha_app = cfg_smtp.get("senha_app", "").strip()
    if not conta or not senha_app:
        raise RuntimeError(
            "Conta Gmail e Senha de App não configuradas.\n"
            "Configure-as na aba 'Configurar E-mails'."
        )

    ok = falhas = 0
    for item in selecionados:
        nome, caminhos = item["nome"], item["caminhos"]
        destino = cfg_emails.get(item["chave"], {})
        para = destino.get("para", "").strip()
        cc = destino.get("cc", "").strip()
        if not para:
            log(f"- Pulado ({nome}): sem e-mail 'Para' configurado.")
            falhas += 1
            continue
        assunto = resolver_template(assunto_tmpl, nome, mes, ano)
        corpo = resolver_template(corpo_tmpl, nome, mes, ano)
        if modo_rascunho and confirmar_fn:
            acao = confirmar_fn(nome, para, cc, assunto, corpo, caminhos)
            if acao == "cancelar":
                log("- Envio cancelado pelo usuário.")
                break
            if acao == "pular":
                log(f"- Pulado ({nome}): cancelado na prévia.")
                falhas += 1
                continue
        try:
            enviar_email_smtp(conta, senha_app, para, cc, assunto, corpo, caminhos)
            log(f"✓ Enviado: {nome} → {para}" + (f" (cc: {cc})" if cc else ""))
            ok += 1
        except Exception as e:
            log(f"✗ ERRO ao enviar e-mail de {nome}: {e}")
            falhas += 1
    return ok, falhas


# ─────────────────────────── Interface (tkinter) ─────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Formatador de Comissões")
        self.geometry("800x720")
        self.resizable(True, True)

        self.cfg = load_config()

        self._build_header()

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True)
        self.nb = nb

        nb.add(self._build_tab_processar(nb), text="Processar")
        nb.add(self._build_tab_emails(nb), text="Configurar E-mails")
        nb.add(self._build_tab_enviar(nb), text="Enviar")

        nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    # ── Cabeçalho: Ano/Mês, compartilhado por todas as abas ────────────────────

    def _build_header(self):
        cfg = self.cfg
        header = ttk.Frame(self, padding=(12, 10, 12, 6))
        header.pack(fill="x")

        ttk.Label(header, text="Ano:").pack(side="left")
        self.var_ano = tk.StringVar(value=str(cfg.get("ano", 2026)))
        cb_ano = ttk.Combobox(header, textvariable=self.var_ano, width=8, state="readonly",
                              values=["2025", "2026", "2027"])
        cb_ano.pack(side="left", padx=(4, 20))

        ttk.Label(header, text="Mês:").pack(side="left")
        self.var_mes = tk.StringVar(value=str(cfg.get("mes", 6)))
        cb_mes = ttk.Combobox(header, textvariable=self.var_mes, width=14, state="readonly",
                              values=[f"{k} - {v}" for k, v in MESES.items()])
        cb_mes.pack(side="left", padx=4)
        # normaliza o valor exibido de mês
        self.var_mes.set(f"{int(cfg.get('mes', 6))} - {MESES[int(cfg.get('mes', 6))]}")

        cb_ano.bind("<<ComboboxSelected>>", self._on_periodo_changed)
        cb_mes.bind("<<ComboboxSelected>>", self._on_periodo_changed)

        ttk.Separator(self).pack(fill="x")

    def _ano_mes_atual(self):
        ano = int(self.var_ano.get())
        mes = int(self.var_mes.get().split(" - ")[0])
        return ano, mes

    def _on_periodo_changed(self, _event=None):
        try:
            ano, mes = self._ano_mes_atual()
        except Exception:
            return
        save_config({"ano": ano, "mes": mes})
        self.cfg.update({"ano": ano, "mes": mes})
        self._atualizar_tab_emails()
        self._atualizar_lista_envio()

    # ── Aba: Processar ────────────────────────────────────────────────────────

    def _build_tab_processar(self, nb):
        frame = ttk.Frame(nb, padding=12)
        cfg = self.cfg
        pad = {"padx": 8, "pady": 6}

        # Pasta Downloads
        self.var_down = tk.StringVar(value=cfg.get("downloads", DEFAULT_DOWNLOADS))
        self._pasta_row(frame, "Pasta de Downloads:", self.var_down, pad)

        # Pasta saída
        self.var_out = tk.StringVar(value=cfg.get("output", DEFAULT_OUTPUT))
        self._pasta_row(frame, "Pasta de saída:", self.var_out, pad)

        # Botão processar
        ttk.Button(frame, text="Processar", command=self.on_processar).pack(**pad)

        # Log
        self.log_widget = scrolledtext.ScrolledText(frame, height=16, wrap="word")
        self.log_widget.pack(fill="both", expand=True, **pad)
        return frame

    def _pasta_row(self, parent, label, var, pad):
        row = ttk.Frame(parent)
        row.pack(fill="x", **pad)
        ttk.Label(row, text=label, width=18).pack(side="left")
        ttk.Entry(row, textvariable=var).pack(side="left", fill="x", expand=True, padx=4)
        ttk.Button(row, text="Procurar…",
                   command=lambda: self._escolher_pasta(var)).pack(side="left")

    def _escolher_pasta(self, var):
        d = filedialog.askdirectory(initialdir=var.get() or os.path.expanduser("~"))
        if d:
            var.set(d)

    def log(self, msg):
        self.log_widget.insert("end", msg + "\n")
        self.log_widget.see("end")
        self.update_idletasks()

    def on_processar(self):
        try:
            ano, mes = self._ano_mes_atual()
        except Exception:
            messagebox.showerror("Erro", "Ano/Mês inválidos.")
            return
        downloads = self.var_down.get().strip()
        output = self.var_out.get().strip()

        save_config({"ano": ano, "mes": mes, "downloads": downloads, "output": output})
        self.cfg.update({"ano": ano, "mes": mes, "downloads": downloads, "output": output})

        self.log_widget.delete("1.0", "end")
        self.log(f"Processando {mes:02d}/{ano}…")
        processar(ano, mes, downloads, output, self.log, self._escolher_duplicado)

    def _escolher_duplicado(self, base, arquivos):
        """Modal para escolher qual arquivo usar quando há duplicados."""
        dlg = tk.Toplevel(self)
        dlg.title("Arquivo duplicado")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, padding=10,
                  text=f"Mais de um arquivo para:\n{base}\n\nQual usar?").pack(anchor="w")
        escolha = tk.StringVar(value=arquivos[0])
        for a in arquivos:
            ttk.Radiobutton(dlg, text=a, value=a, variable=escolha).pack(anchor="w", padx=16)
        resultado = {"valor": None}

        botoes = ttk.Frame(dlg, padding=10)
        botoes.pack(fill="x")

        def ok():
            resultado["valor"] = escolha.get()
            dlg.destroy()

        def pular():
            resultado["valor"] = None
            dlg.destroy()

        ttk.Button(botoes, text="Usar este", command=ok).pack(side="right", padx=4)
        ttk.Button(botoes, text="Pular", command=pular).pack(side="right")
        self.wait_window(dlg)
        return resultado["valor"]

    # ── Aba: Configurar E-mails ───────────────────────────────────────────────

    def _build_tab_emails(self, nb):
        frame = ttk.Frame(nb, padding=12)

        # Seção conta Gmail
        smtp_frame = ttk.LabelFrame(frame, text="Conta Gmail para envio", padding=8)
        smtp_frame.pack(fill="x", pady=(0, 12))

        cfg_smtp = self.cfg.get("smtp", {})
        ttk.Label(smtp_frame, text="Conta (e-mail):").grid(
            row=0, column=0, sticky="w", padx=4, pady=4)
        self.var_smtp_conta = tk.StringVar(value=cfg_smtp.get("conta", ""))
        ttk.Entry(smtp_frame, textvariable=self.var_smtp_conta, width=44).grid(
            row=0, column=1, sticky="we", padx=4)

        ttk.Label(smtp_frame, text="Senha de App:").grid(
            row=1, column=0, sticky="w", padx=4, pady=4)
        self.var_smtp_senha = tk.StringVar(value=cfg_smtp.get("senha_app", ""))
        ttk.Entry(smtp_frame, textvariable=self.var_smtp_senha, show="*", width=44).grid(
            row=1, column=1, sticky="we", padx=4)

        ttk.Label(
            smtp_frame,
            text="Crie em: myaccount.google.com > Segurança > Senhas de app",
            font=("", 8, "italic"),
            foreground="gray",
        ).grid(row=2, column=0, columnspan=2, sticky="w", padx=4, pady=(2, 4))

        ttk.Button(smtp_frame, text="Salvar conta Gmail", command=self._salvar_smtp).grid(
            row=3, column=0, columnspan=2, sticky="w", padx=4, pady=(4, 0))
        smtp_frame.columnconfigure(1, weight=1)

        ttk.Separator(frame).pack(fill="x", pady=(0, 8))

        ttk.Label(
            frame,
            text="Destinatários por equipe: e-mails separados por ponto e vírgula (;)",
            font=("", 9, "italic"),
        ).pack(anchor="w", pady=(0, 8))

        self.emails_container = ttk.Frame(frame)
        self.emails_container.pack(fill="both", expand=True)
        self.email_vars = {}

        ttk.Button(frame, text="Salvar destinatários", command=self._salvar_emails).pack(
            anchor="w", pady=12)
        return frame

    def _atualizar_tab_emails(self):
        for child in self.emails_container.winfo_children():
            child.destroy()

        try:
            ano, mes = self._ano_mes_atual()
        except Exception:
            return

        downloads = self.var_down.get().strip()
        output = self.var_out.get().strip()
        itens = grupos_do_mes(downloads, output, ano, mes)
        cfg_emails = self.cfg.get("emails", {})

        ttk.Label(self.emails_container, text="Equipe", width=22, font=("", 9, "bold")).grid(
            row=0, column=0, sticky="w", padx=4, pady=4)
        ttk.Label(self.emails_container, text="Para (líder)", font=("", 9, "bold")).grid(
            row=0, column=1, sticky="w", padx=4, pady=4)
        ttk.Label(self.emails_container, text="Cc", font=("", 9, "bold")).grid(
            row=0, column=2, sticky="w", padx=4, pady=4)

        if not itens:
            ttk.Label(
                self.emails_container,
                text="Nenhuma equipe encontrada para este Ano/Mês (baixe ou processe primeiro).",
            ).grid(row=1, column=0, columnspan=3, sticky="w", padx=4, pady=8)

        self.email_vars = {}
        for i, (chave, nome, _slugs) in enumerate(itens, start=1):
            atual = cfg_emails.get(chave, {})
            var_para = tk.StringVar(value=atual.get("para", ""))
            var_cc = tk.StringVar(value=atual.get("cc", ""))
            self.email_vars[chave] = {"para": var_para, "cc": var_cc}

            ttk.Label(self.emails_container, text=nome, width=22).grid(
                row=i, column=0, sticky="w", padx=4, pady=2)
            ttk.Entry(self.emails_container, textvariable=var_para).grid(
                row=i, column=1, sticky="we", padx=4, pady=2)
            ttk.Entry(self.emails_container, textvariable=var_cc).grid(
                row=i, column=2, sticky="we", padx=4, pady=2)

        self.emails_container.columnconfigure(1, weight=1)
        self.emails_container.columnconfigure(2, weight=1)

    def _salvar_emails(self):
        # começa do que já está salvo, para não apagar equipes de outros meses
        emails = dict(self.cfg.get("emails", {}))
        for chave, v in self.email_vars.items():
            emails[chave] = {"para": v["para"].get().strip(), "cc": v["cc"].get().strip()}
        self.cfg["emails"] = emails
        save_config({"emails": emails})
        messagebox.showinfo("Destinatários", "Lista de e-mails salva.")

    def _salvar_smtp(self):
        smtp = {
            "conta": self.var_smtp_conta.get().strip(),
            "senha_app": self.var_smtp_senha.get().strip(),
        }
        self.cfg["smtp"] = smtp
        save_config({"smtp": smtp})
        messagebox.showinfo("Conta Gmail", "Conta Gmail salva.")

    # ── Aba: Enviar ────────────────────────────────────────────────────────────

    def _build_tab_enviar(self, nb):
        frame = ttk.Frame(nb, padding=12)

        ttk.Button(frame, text="↻ Atualizar lista", command=self._atualizar_lista_envio).pack(
            anchor="w", pady=(0, 8))

        self.lista_frame = ttk.Frame(frame)
        self.lista_frame.pack(fill="x", pady=(0, 8))
        self.enviar_vars = {}

        ttk.Label(frame, text="Assunto:").pack(anchor="w")
        cfg_envio = self.cfg.get("envio", {})
        self.var_assunto = tk.StringVar(value=cfg_envio.get("assunto", DEFAULT_ASSUNTO))
        ttk.Entry(frame, textvariable=self.var_assunto).pack(fill="x", pady=(0, 8))

        ttk.Label(frame, text="Corpo (use {equipe}, {mes}, {ano}):").pack(anchor="w")
        self.txt_corpo = scrolledtext.ScrolledText(frame, height=8, wrap="word")
        self.txt_corpo.insert("1.0", cfg_envio.get("corpo", DEFAULT_CORPO))
        self.txt_corpo.pack(fill="x", pady=(0, 8))

        modo_row = ttk.Frame(frame)
        modo_row.pack(fill="x", pady=(0, 8))
        self.var_modo = tk.StringVar(value=cfg_envio.get("modo", "rascunho"))
        ttk.Radiobutton(modo_row, text="Prévia antes de enviar (confirmar um a um)",
                        variable=self.var_modo, value="rascunho").pack(side="left", padx=(0, 16))
        ttk.Radiobutton(modo_row, text="Enviar direto",
                        variable=self.var_modo, value="direto").pack(side="left")

        botoes_row = ttk.Frame(frame)
        botoes_row.pack(fill="x", pady=(0, 8))
        ttk.Button(botoes_row, text="Salvar padrão de assunto/corpo",
                   command=self._salvar_padrao_envio).pack(side="left", padx=(0, 8))
        ttk.Button(botoes_row, text="Executar envio",
                   command=self._executar_envio_click).pack(side="left")

        self.log_widget_envio = scrolledtext.ScrolledText(frame, height=10, wrap="word")
        self.log_widget_envio.pack(fill="both", expand=True)
        return frame

    def _salvar_padrao_envio(self):
        envio = dict(self.cfg.get("envio", {}))
        envio.update({
            "modo": self.var_modo.get(),
            "assunto": self.var_assunto.get(),
            "corpo": self.txt_corpo.get("1.0", "end").strip(),
        })
        self.cfg["envio"] = envio
        save_config({"envio": envio})
        messagebox.showinfo("Padrão de e-mail", "Assunto e corpo padrão salvos.")

    def _log_envio(self, msg):
        self.log_widget_envio.insert("end", msg + "\n")
        self.log_widget_envio.see("end")
        self.update_idletasks()

    def _on_tab_changed(self, _event):
        try:
            aba_atual = self.nb.tab(self.nb.select(), "text")
        except Exception:
            return
        if aba_atual == "Enviar":
            self._atualizar_lista_envio()
        elif aba_atual == "Configurar E-mails":
            self._atualizar_tab_emails()

    def _atualizar_lista_envio(self):
        for child in self.lista_frame.winfo_children():
            child.destroy()
        self.enviar_vars = {}

        try:
            ano, mes = self._ano_mes_atual()
        except Exception:
            messagebox.showerror("Erro", "Ano/Mês inválidos.")
            return

        output_dir = self.var_out.get().strip()
        cfg_emails = self.cfg.get("emails", {})
        grupos = montar_grupos_envio(output_dir, ano, mes)

        if not grupos:
            ttk.Label(
                self.lista_frame,
                text="Nenhum arquivo encontrado para este Ano/Mês na pasta de saída.",
            ).pack(anchor="w")
            return

        for grupo in grupos:
            chave, nome, arquivos = grupo["chave"], grupo["nome"], grupo["arquivos"]
            caminhos = [a["caminho"] for a in arquivos]
            nomes_arquivos = ", ".join(os.path.basename(c) for c in caminhos)
            para = cfg_emails.get(chave, {}).get("para", "").strip()
            var_check = tk.BooleanVar(value=bool(para))

            texto = f"{nome} ({nomes_arquivos})"
            if not para:
                texto += "  (sem e-mail configurado, veja a aba Configurar E-mails)"

            ttk.Checkbutton(
                self.lista_frame, text=texto, variable=var_check,
                state=("normal" if para else "disabled"),
            ).pack(anchor="w", pady=1)

            self.enviar_vars[chave] = {"check": var_check, "caminhos": caminhos, "nome": nome}

    def _executar_envio_click(self):
        try:
            ano, mes = self._ano_mes_atual()
        except Exception:
            messagebox.showerror("Erro", "Ano/Mês inválidos.")
            return

        selecionados = [
            {"chave": chave, "nome": v["nome"], "caminhos": v["caminhos"]}
            for chave, v in self.enviar_vars.items() if v["check"].get()
        ]
        if not selecionados:
            messagebox.showwarning("Enviar", "Nenhum arquivo selecionado.")
            return

        assunto_tmpl = self.var_assunto.get()
        corpo_tmpl = self.txt_corpo.get("1.0", "end").strip()
        modo = self.var_modo.get()
        cfg_emails = self.cfg.get("emails", {})
        cfg_smtp = self.cfg.get("smtp", {})

        self.log_widget_envio.delete("1.0", "end")
        try:
            ok, falhas = executar_envio(
                selecionados, cfg_emails, cfg_smtp, assunto_tmpl, corpo_tmpl,
                ano, mes, modo == "rascunho", self._log_envio,
                confirmar_fn=self._confirmar_envio_modal if modo == "rascunho" else None,
            )
        except RuntimeError as e:
            messagebox.showerror("Gmail", str(e))
            return

        resumo = f"{ok} e-mail(s) enviado(s)."
        if falhas:
            resumo += f" {falhas} falharam ou foram pulados (veja o log)."
        messagebox.showinfo("Enviar", resumo)


    def _confirmar_envio_modal(self, nome, para, cc, assunto, corpo, caminhos):
        """Mostra prévia do e-mail e pede confirmação. Retorna 'enviar', 'pular' ou 'cancelar'."""
        dlg = tk.Toplevel(self)
        dlg.title(f"Prévia: {nome}")
        dlg.transient(self)
        dlg.grab_set()
        dlg.geometry("580x460")

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill="both", expand=True)

        def _linha(label, valor):
            ttk.Label(frm, text=label, font=("", 9, "bold")).pack(anchor="w")
            ttk.Label(frm, text=valor or "(vazio)", wraplength=540, justify="left").pack(
                anchor="w", padx=8, pady=(0, 8))

        _linha("Para:", para)
        if cc:
            _linha("Cc:", cc)
        _linha("Assunto:", assunto)
        _linha("Anexos:", ", ".join(os.path.basename(c) for c in caminhos))
        ttk.Label(frm, text="Corpo:", font=("", 9, "bold")).pack(anchor="w")
        txt = scrolledtext.ScrolledText(frm, height=8, wrap="word")
        txt.insert("1.0", corpo)
        txt.config(state="disabled")
        txt.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        resultado = {"valor": "pular"}

        botoes = ttk.Frame(dlg, padding=(12, 0, 12, 12))
        botoes.pack(fill="x")

        def _enviar():
            resultado["valor"] = "enviar"
            dlg.destroy()

        def _pular():
            resultado["valor"] = "pular"
            dlg.destroy()

        def _cancelar():
            resultado["valor"] = "cancelar"
            dlg.destroy()

        ttk.Button(botoes, text="Enviar", command=_enviar).pack(side="right", padx=4)
        ttk.Button(botoes, text="Pular", command=_pular).pack(side="right", padx=4)
        ttk.Button(botoes, text="Cancelar tudo", command=_cancelar).pack(side="left")

        self.wait_window(dlg)
        return resultado["valor"]


if __name__ == "__main__":
    App().mainloop()
